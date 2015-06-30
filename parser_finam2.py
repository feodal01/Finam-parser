import urllib.request
from bs4 import BeautifulSoup
import datetime
import schedule
import time
from openpyxl import Workbook 
from openpyxl import load_workbook

URL = 'https://www.finam.ru/'
leaders = []    
outsaiders = []

def what_time_is_it():
    x = str(datetime.datetime.now()).split(' ')[1].split('.')[0].split(':')
    return str(x[0]) + '-' + str(x[1])

def job():
    #print( str(what_time_is_it()))
    print('парсим по расписанию'+ '('+ str(what_time_is_it()) + ')' ) 
    def get_html(URL):  #прочитали URL  
        response = urllib.request.urlopen(URL)
        return response.read()

    def parse(html):
        soup = BeautifulSoup(html) #переводим в html формат
        table = soup.find(class_='home-page-leaders-table-rus') #ищем класс (там табличка находится наша)
        rows = table.find_all('tr') #ищем все с тегом тр/ если понадобится убрать какую то из строк, то надо срез: [1:]
        global leaders
        global outsaiders 
         
        for row in rows:
            cols = row.find_all('td')
            
            leaders.append({
                'title': cols[0].a.text.split()[0],
                'price': float((cols[1].span.text).replace(',','.'))
                })
                
            outsaiders.append({
                'title': cols[3].a.text.split()[0],
                'price': float((cols[4].span.text).replace(',','.'))      
                })        
        return leaders, outsaiders    
    
    def export_excel(leader, outsider, filename):

        def write(): #эта функция непосредственно осуществляет запись
            worksheet = workbook.create_sheet() #создаем новый лист х2
            worksheet1 = workbook.create_sheet()
            x = str(datetime.datetime.now()).split(' ')[1].split('.')[0].split(':') #определяем сколько времени
            worksheet.title = '%s' % ('leader-' + str(x[0]) + '-' + str(x[1])) #обзываем вкладку как лидер-время
            worksheet1.title = '%s' % ('outsider-' + str(x[0]) + '-' + str(x[1])) #то же для аутсайдера
     
            field_names = ('Бумага','Изменение') #задаем названия столбцов         
            for i, field in enumerate(field_names): #энумерэйт заменяет счетчик
                worksheet.cell(column = i+1, row = 1).value = '%s' % field #записываем в ячейку название столбца
                worksheet1.cell(column = i+1, row = 1).value = '%s' % field     
     
            fields = ('title','price')   

            for row, leader in enumerate(leaders, start=1):  #записываем во вкладку лидеров
                for col, field in enumerate(fields):
                    worksheet.cell(column = col+1, row = row+1).value = leader[field] #записываем в ячейку данные
     
            for row, outsaider in enumerate(outsaiders, start=1):  #записываем во вкладку аутсайдеров
                for col, field in enumerate(fields):
                    worksheet1.cell(column = col+1, row = row+1).value = outsaider[field] #записываем в ячейку данные
            print('запись в файл завершена')
        try: #определяем существует ли файл за сегодня. если да то дописываем, если нет то  создаем
            workbook = load_workbook(filename)
            print('книгу загрузили, дополняем новыми лиcтами!')            
            write()
            workbook.save(filename)
        except:
            print('не удалось загрузить книгу, делаем новую!')
            workbook = Workbook() #создаем книгу
            write()
            workbook.save(filename)         
        
    """ ЗАПУСКАЕМ ПАРСИНГ И СОХРАНЕНИЕ РЕЗУЛЬТАТОВ!"""
    try:
        parse(get_html(URL))
        export_excel(leaders, outsaiders, str(datetime.datetime.now()).split(' ')[0] + '.xlsx')
    except:
        print('Не удалось запарсить страницу')
    global leaders
    global outsaiders 
    leaders, outsaiders = [], []
    print('парсинг завершен')
    
def main():
    schedule.every().day.at("13:10").do(job)
    schedule.every().day.at("16:30").do(job) #тут мы устанавливаем когда будем запускать функцию джоб
    while True:
        schedule.run_pending()
        time.sleep(1)

if __name__ == '__main__':
    main()

